"""
author: Basel Alyafi
year: 2019
Erasmus Mundus in MEdical Imaging and Applications 2017-2019
Master's Thesis Project
"""
import shutil
from glob import glob

import matplotlib
import matplotlib.pyplot as plt
import seaborn

import torch
import torch.nn as nn
from torchvision.datasets import ImageFolder  # to create the dataset
from torch.utils.data import DataLoader, Subset  # to create the needed real_loaders
from torch.utils.data.sampler import SubsetRandomSampler, WeightedRandomSampler, SequentialSampler # to split the dataset
from torchvision import transforms
from torchvision.utils import make_grid

from openpyxl import Workbook, load_workbook
from sklearn.utils.multiclass import unique_labels
from sklearn.metrics import precision_recall_fscore_support, roc_curve, average_precision_score, roc_auc_score, confusion_matrix
from sklearn.model_selection import KFold
from sklearn.manifold import TSNE

import numpy as np
from scikitplot.metrics import plot_precision_recall
import copy, os

# to modify docstring
# noinspection SpellCheckingInspection
from Evaluation.classes import Classifier
from Gans.DCGAN import Generator
from termcolor import cprint
import time

from skimage import io, color

def data_loaders(path, train_share, val_share, shuffle, batch_size, n_splits = 10, transformations=None): #DOC OK
    """
    This function is aimed at reading the dataset, splitting into train/val/test, and returns the data loaders needed, the train percentage
    is 1 - val_share. It is designed especially for imbalanced binary datasets.

    Params:
    ------

    path: string
        the path to the dataset root folder, e.g., path/cat/x.ext
    train_share: list of float/int
        if <=1: the training percentage for each class [train0, train1], this size is relative to (n_splits-1) * size(dataset)/(n_splits)
        if >1: the number of training images, should not exceed number of images in (n_splits-1) * size(dataset)/(n_splits)
    val_share: list of float/int numbers
        if <1: the validation percentage for each class [val0, val1], this size is relative to (n_splits-1) * size(dataset)/(n_splits)
        if >1: the number of validation images
        if val_share and train_share are integer: val_share + train_share <=  (n_splits-1) * size(dataset)/(n_splits)
    shuffle: boolean
        to shuffle the names before sampling
    batch_size: int
        number of images per batch
    n_splits: int
        number of cross validation folds.
    transformations: Tensor Transforms
        the transforms to apply on images when creating batches
    Returns:
    -------
    real_loaders: list of dictionaries
        training, validation, testing data loaders are included in each dictionary in the list. The length of the list is equal to n_splits.

    Notes:
    -----
    train_share values and val_share should be consistent, meaning both values in each list should be of the same data type (int/float).
    if both lists are float, then the following should hold:
        train_share[i] + val_share[i] <= 1; i= 0,1
    """

    # compose the needed transforms if not passed as a parameter
    if not transformations:
        train_transform = transforms.Compose([
                                            transforms.Grayscale(),
                                            transforms.Resize([128, 128]),
                                            transforms.ToTensor()
                                        ])
    else:
        train_transform = transformations

    # creating the dataset
    dataset = ImageFolder(path, transform=train_transform)

    # distinguishing between the two classes
    first_indices = [idx for idx in range(len(dataset.imgs)) if dataset.imgs[idx][1] == 0]
    first_indices = np.reshape(first_indices, [len(first_indices),1])

    second_indices = [idx for idx in range(len(dataset.imgs)) if dataset.imgs[idx][1] == 1]
    second_indices = np.reshape(second_indices, [len(second_indices), 1])

    # shuffle the dataset
    np.random.seed(119)
    np.random.shuffle(first_indices)
    np.random.shuffle(second_indices)

    folds = KFold(n_splits=n_splits, shuffle=shuffle, random_state=119)#was 109 for mass lesions

    # create a list for the samplers
    samplers = []
    fold_idx=0

    # loop over all the folds of both classes
    for (train_0idx, test_0indices), (train_1idx, test_1indices) in zip(folds.split(first_indices), folds.split(second_indices)):

        # extract the needed share of the fold only
        train_0share = extract_indices(train_0idx, partition=train_share[0], start_idx=0, shuffle_output=True)
        val_0share   = extract_indices(train_0idx, partition=val_share[0], start_idx=len(train_0share), shuffle_output=True)

        train_1share = extract_indices(train_1idx, partition=train_share[1], start_idx=0, shuffle_output=True)
        val_1share   = extract_indices(train_1idx, partition=val_share[1], start_idx=len(train_1share), shuffle_output=True)

        # select the corresponding first and second class indices
        train_0indices, val_0indices, test_0indices = first_indices[train_0share], first_indices[val_0share], first_indices[test_0indices]
        train_1indices, val_1indices, test_1indices = second_indices[train_1share], second_indices[val_1share], second_indices[test_1indices]

        # combine first class and second class indices for train, val, and test.
        train_indices = np.append(train_0indices, train_1indices)
        val_indices = np.append(val_0indices, val_1indices)
        test_indices = np.append(test_0indices, test_1indices)

        # create a dictionary for training/validation/testing indices
        sampler = dict()
        sampler['train'] = SubsetRandomSampler(train_indices)
        sampler['val'] = SubsetRandomSampler(val_indices)
        sampler['test'] = SubsetRandomSampler(test_indices)

        # append this fold dictionary to the list of all folds dictionaries
        samplers.append(sampler)
        fold_idx += 1

    # create a dictionary for the real_loaders ['train', 'val', 'test']
    loaders = []

    # loop over all folds and create a dictionary of dataloaders for each using corresponding samplers.
    for idx in range(n_splits):
        loader= dict()
        loader['train'] = DataLoader(dataset, batch_size=batch_size, sampler=samplers[idx]['train'], num_workers=8)
        loader['val'] = DataLoader(dataset, batch_size=batch_size, sampler=samplers[idx]['val'], num_workers=8)
        loader['test'] = DataLoader(dataset, batch_size=batch_size, sampler=samplers[idx]['test'], num_workers=8)# shuffle does not work with samplers

        # append the dictionary to the all folds list of loaders
        loaders.append(loader)
    return loaders

def data_loader(path, batch_size, transforms, shuffle):
    """
    Params:
    -------
    path: string
        the path were data folders are located, each class should be in a separate folder
    batch_size: int
        the size of the batch the loader returns every iteration.
    transforms: torchvision.transforms
        the transforms to apply on input images
    shuffle: bool
        whether to shuffle images every epoch or not.

    Returns:
    data_loader: torch.utils.data.dataloader.DataLoader
        the data loader to load a batch every times

    -------
    """
    dataset = ImageFolder(path, transforms)
    loader = DataLoader(dataset, batch_size=batch_size, shuffle=shuffle, num_workers=4)
    return loader

# noinspection SpellCheckingInspection,PyUnresolvedReferences
def train_model(model, data_loader, optimizer, scheduler, epoch_num, print_every, validate_every, pos_weight=None, save=True, initial_f1=0, plot=True): #DOC OK
    """
    This helper is aimed at training a specific model for a specific number of epochs.

    Params:
    -------
    model: nn.Module
        the model to be trained
    data_loader: dict
        the data real_loaders dictionary containing three objects of torch.utils.data.DataLoader
    optimizer: optim
        the optimizer to use in the training process
    scheduler: optimizer.lr_scheduler
        learning rate scheduler
    epoch_num: int
        number of epochs to train the model
    print_every: int
        number of batches after which the loss is shown
    validate_every: int
        number of batches after which the validation step is done, if -1 no validation will be done

    pos_weight: int/float
        if provided, BCEWithLogitsLoss is used with this value as the weight for the positive class
    save: boolean
        whether to save_model the resulted model or not
    initial_f1: int
        the initial F1 score, used when looping over multiple datasets training the same model to save the best model with highest F1 among all datasets.
    plot: boolean
        whether to plot the training nad validation losses or not.
    Returns:
    -------
    tuple(model, f1): (nn.Module, float)
        the trained model and the best f1 score achieved
     """

    # set the mode to training
    model.train(mode=True)

    # initialize the stopwatch
    start = time.time()

    # DEBUGGING
    print('learning rate is {}'.format(scheduler.get_lr()))

    # initialization
    losses = []
    val_losses = []
    val_loss=1e6
    epoch=0
    f1 = initial_f1

    # decide whether cpu or gpu
    device = Classifier.device

    correct_preds, cur_loss = torch.zeros(1), 0
    num_batches = len(data_loader['train'])

    # for all epochs
    for epoch in range(epoch_num):

        cur_loss = 0
        correct_preds = 0

        # for all batches
        for batch_num, (imgs, labels) in enumerate(data_loader['train']):

            # move images and labels to the correct memory (cpu, gpu)
            imgs, labels = imgs.to(device), labels.to(device).float()

            # starting calculating gradients every batch from scratch
            optimizer.zero_grad()
            model.zero_grad()

            # start training
            with torch.set_grad_enabled(True):
                out = model(imgs).view(-1)  #get the model output
                probs = torch.sigmoid(out)  # get probabilities
                pred = torch.gt(probs, 0.5).float() # put the threshold for 1 classification

                if not(isinstance(pos_weight, int) or isinstance(pos_weight, float)):
                    pos_weight = 1
                loss = auto_weight_BCEWithLogitsLoss(output=out, targets=labels, pos_weight=pos_weight)

                loss.backward()
                optimizer.step()

            # count the correct predicted labels over all batches in the epoch
            correct_preds += torch.sum(pred == labels.data)
            # accummulate the loss
            cur_loss += loss.item() * imgs.size(0)

            # append training losses
            if (batch_num % print_every == 0)| (batch_num==num_batches-1):
                losses.append(loss.item())
                print('[{}/{}] [{}/{}] loss= {:.5f}'.format(epoch, epoch_num - 1, batch_num, len(data_loader['train']), loss.item()))

            # validation
            if (validate_every != -1) and ((batch_num % validate_every == 0)| (batch_num==num_batches-1)):
                print('validating...')
                _, cur_val_loss, cur_f1 = predict(model, data_loader['val'], pos_weight)
                val_losses.append(cur_val_loss)

                # save the model if required with the best configuration
                if cur_val_loss < val_loss:
                    val_loss = cur_val_loss

                if save & (f1 < cur_f1[1]):
                    f1 = cur_f1[1]
                    abs_path = os.path.join(model.path, model.id)

                    if not os.path.exists(model.path):
                        os.makedirs(model.path, exist_ok=True)
                    torch.save(model.state_dict(), abs_path )
                    print('-'*40 + 'model '+ os.path.split(model.path)[1] + ' saved' + '-'*40)

        scheduler.step()
        # normalize by dividing by the total number of cases
        acc = correct_preds.double() / (data_loader['train'].sampler.num_samples) #was .indices.size
        cur_loss /= (data_loader['train'].sampler.num_samples)

        print(str.upper('train') + '-' * 20 + 'Epoch {} Finished, Train Acc {:.4f} Train Loss {:.4f}'.format(epoch, acc.item(), cur_loss))
    end = time.time()
    print('Time Elapsed is {} Min'.format((end-start)/60))

    # plotting the progress
    if plot:
        plt.figure()
        avg_loss = np.convolve(losses, np.ones(25) / 25, mode='same')
        avg_loss[:13]= losses[:13]
        avg_loss[-13:] = losses[-13:]

        plt.plot(losses, label='train loss')
        plt.plot(avg_loss, label='average train loss', linewidth=2)
        plt.legend()

        plt.xlabel(' iterations' + ' X '+str(print_every))
        plt.ylabel('BCE loss')
        plt.title('Training Loss')
        plt.savefig(os.path.join(model.path,'losses'))

        plt.figure()

        avg_val_loss = np.convolve(val_losses, np.ones(25) / 25, mode='same')
        avg_val_loss[:13] = val_losses[:13]
        avg_val_loss[-13:] = val_losses[-13:]

        plt.plot(val_losses, label='validation loss')
        plt.plot(avg_val_loss, label='average val loss', linewidth=2)
        plt.legend()

        plt.xlabel('validation iterations' + 'every '+str(validate_every))
        plt.ylabel('BCE loss')
        plt.title('Val Loss')
        plt.savefig(os.path.join(model.path,'avg_val_losses'))

        plt.show()

    # load the best saved model
    if save:
        model.load_state_dict(torch.load(abs_path))
        cprint('\nBest Val Results:', 'yellow', attrs=['bold'])
        predict(model, data_loader['val'], pos_weight)

    return model, f1


def initialize_weights(layer):#DOC OK
    """
    This function is aimed at initializing a model using random values taken from norma distribution with Xavier standard deviation .

    Params:
    -------
    layer: torch.nn
        a layer un a torch.nn.Module
    Return:
    -------
    void, it normalizes the layer weights using xavier normalization.
    """
    class_name = layer.__class__.__name__
    if type(layer) == nn.Conv2d:
        nn.init.xavier_normal_(layer.weight.data)
    elif class_name.find('BatchNorm') != -1:
        nn.init.normal_(layer.weight.data, 1.0, 0.02)
        nn.init.constant_(layer.bias.data, 0)


def predict(model, data_loader, pos_loss_weight, verbose=0, tag='', xlfile_name=None, prefix=None): #DOC OK
    """
    This function is used to evaluate a binary classifier.

    Params:
    ------

    model: torch.nn.Module
        the model to be used for prediction
    data_loader: torch.utils.data.DataLoader
        the test data real_loaders
    pos_loss_weight: int/float
        a number to be passed as a the weight for the positive class for BCEWithLogitsLoss

    verbose: int
        if 0 : only metrics will be printed

        if 1 : verbose 0 + plotting precision-recall curve

        if 2 : verbose 1 + visualize 40 images of wrong and correct predictions.

        if 3: verbose 2 + visualize the confusion matrix

    tag: string
        a string to include in the save figure name.

    xlfile_name: str
        the path where the results Excel file is located.
    prefix: list
        the values to append results to, usually are the parameters used to get these results.
    Returns:
    -------
        tuple(predicted labels, the loss value, F1 score)
    Notes:
    ------
    Sigmoid is applied on the output of the model, so no need to include it in the model.

    """

    # set the device
    device = Classifier.device

    matplotlib.use('Agg') #TODO disable when you want to show figures
    mode = model.training
    # set the mode to evaluation
    model.eval()

    np_pred_labels = np.array([])
    np_true_labels = np.array([])
    np_probs = np.array([])

    # for all batches
    for idx ,(imgs, labels) in enumerate(data_loader):
        imgs, labels = imgs.to(device), labels.to(device).float()

        # start testing with no auto grad
        with torch.set_grad_enabled(False):
            output = model(imgs).view(-1) #get the output batch
            probs = torch.sigmoid(output)

            preds = torch.gt(probs, 0.5).float()
            loss = auto_weight_BCEWithLogitsLoss(output, labels, pos_loss_weight)

        # stack all the true labels, predicted labels, and probabilities for all batches together
        np_true_labels = np.hstack((np_true_labels, labels.data.cpu().numpy())).astype(np.int)
        np_pred_labels = np.hstack((np_pred_labels, preds.data.cpu().numpy())).astype(np.int)
        np_probs = np.hstack((np_probs, probs.data.cpu().numpy()))

    if verbose >= 0: #print metrics
        Pr, R, F1, S = precision_recall_fscore_support(y_true=np_true_labels, y_pred=np_pred_labels, average=None)
        average_precision = average_precision_score(y_true=np_true_labels, y_score=np_probs)

        print('-' * 10 + '(class 0): ' + ' Loss= {:.5f} Precision = {:.4f} Recall = {:.4f} F1 = {:.4f} Support = {}'.format(loss.item(), Pr[0], R[0], F1[0], S[0]))
        print('-'*10+ '(class 1): ' + ' Loss= {:.5f} Precision = {:.4f} Recall = {:.4f} F1 = {:.4f} Support = {}'.format(loss.item(), Pr[1], R[1], F1[1], S[1]))
        print('\naverage precision {:.4f}'.format(average_precision)) # area under precision-recall curve can be useful here

        if verbose >= 1: #plot precision recall curve
            two_class_probs = np.stack((1-np_probs, np_probs), axis=1) #stack positive class and negative class probabilities
            plot_precision_recall(y_true=np_true_labels, y_probas= two_class_probs) # plot precision recall curve

            fpr, tpr, thresholds = roc_curve(y_true= np_true_labels, y_score= np_probs, pos_label=1)
            roc_auc = roc_auc_score(y_true=np_true_labels, y_score=np_probs)
            plt.figure()
            plt.plot(fpr, tpr)
            plt.xlabel('FPR'), plt.ylabel('TPR'), plt.title('ROC Curve for model\n' +  os.path.split(model.path)[1]), plt.legend(['auc = {:.4f}'.format(roc_auc)])
            plt.savefig(os.path.join(model.path, model.id + tag + '_ROC' ))
            if verbose >= 2: #show some images (20 correctly predicted and 20 mistakenly predicted)
                visualize_model(num_images=[20,20], data_loader=data_loader, true_labels=np_true_labels, pred_labels=np_pred_labels ,classes=['Tis', 'Les'])

                if verbose==3:
                    plot_confusion_matrix(y_true=np_true_labels, y_pred=np_pred_labels, classes=np.array(['Tis', 'Les']))
                    plt.savefig(os.path.join(model.path, model.id + tag+ '_confusion'))
                    plt.show()

    #return the model to the mode it came with
    model.train(mode)

    # if there is an excel file name, write the results there
    if xlfile_name:
        print('writing metrics to the excel file...')
        write_xl(xlfile_name, prefix+[Pr[1], R[1], F1[1], roc_auc])

    return np_pred_labels, loss.item(), F1


def auto_weight_BCEWithLogitsLoss(output, targets, pos_weight=None): # DOC OK
    """
    This function is used to calculate automatically the positive weight needed for BCEWithLogitsLoss

    Params
    ------
    output: list(Tensor)
        the predicted probabilities
    targets: list(Tensor), same size of output
        the target labels
    pos_weight: int/float
        if not None it will be used as the positive weight in BCEWithLogitsLoss

    Returns:
    --------
    returns an instance of the loss function with the parameters passed.
    """
    #detect unique classes with counts
    unique_labels, counts = np.unique(targets.data.cpu().numpy(), return_counts=True)

    # compute the weights using counts (the higher the count, the lower the weight)
    if len(counts)>1:
        weights = np.array([np.sum(counts)/count - 1 for count in counts])
        weights = torch.tensor(np.min(weights)*5)
    else:
        weights=None

    # if pos_weight is not None: #use the passed value if any
    if isinstance(pos_weight, int) or isinstance(pos_weight, float):
        weights = torch.tensor(pos_weight)
    else:
        raise ValueError('positive weight should be either integer or float not {}'.format(type(pos_weight)))

    loss_fun = nn.BCEWithLogitsLoss(pos_weight=weights)#instantiate an object of the loss function
    # print('weights are: ', weights, 'counts: ', counts)
    return loss_fun(output, targets)

def visualize_model(num_images, data_loader, true_labels, pred_labels, classes):# DOC OK
    """
    This function is used to show a specified number of truly + wrongly classified images with corresponding predicted + groundtruth labels

    Params:
    -------
    num_images: list(int) of length 2
        the first number is the maximum requested number of truly classified images,
        the second number is the maximum requested number of wrongly classified images.
    data_loader: torch.utils.data.DataLoader
        the loader used to load images
    true_labels: list(Tensor)
        groundtruth labels
    pred_labels: list(Tensor)
        predicted labels
    classes: list(str), length is number of classes
        classes names, sorted accordingly

    Return:
    -------
    void

    """
    # create a new figure
    plt.figure()

    #initialize counts
    correct_count, mistake_count = 0, 0
    fig_index = 0
    idx = 0

    #loop over all images untill all conditions satisfied or run out of images
    for img_batch, label in data_loader:
        for img in img_batch.data.cpu().numpy():

            #whether to show more correct/wrong classifications
            show_correct = (true_labels[idx] == pred_labels[idx]) and (correct_count < num_images[0])
            show_mistake = (true_labels[idx] != pred_labels[idx]) and (mistake_count < num_images[1])

            if show_correct | show_mistake :   #if there is something to show
                #increase counts correspondingly
                if show_correct:
                    correct_count += 1
                if show_mistake:
                    mistake_count += 1

                # create a figure with 4 rows
                axis = plt.subplot(4, np.round(np.sum(num_images) / 4), fig_index + 1)
                axis.axis('off')

                # set the title to show predicted and groundtruth labels
                axis.set_title('pred={},\n GT={}'.format(classes[pred_labels[idx]], classes[true_labels[idx]]))
                plt.imshow(img[0,:,:], cmap='gray') #
                fig_index += 1

            # stop when reaching the requested number of cases
            if (correct_count == num_images[0]) and (mistake_count == num_images[1]):
                break
            idx += 1
    plt.show()

def set_parameter_requires_grad(model, feature_extracting): #DOC Ok
    """
    this function can be used to set the requires_grad of all parameters in all layers to one value.
    It can be used to freeze all the layers in a model in order to use it for transfer learning.


    Params:
    ------
    model: nn.Module
        the model to set requires_grad in to False

    feature_extracting: bool
        whether to set the requires_grad to true or false

    Returns:
    -------
    void
    """

    if feature_extracting:
        for param in model.parameters():
            param.requires_grad = False


def extract_indices(indices_arr, partition, start_idx=0, shuffle_output=False):#DOC OK
    """
    this function is used to extract a part of a sequence of indices knowing the starting point and the end point or size.

    Params:
    -------
    indices_arr: list(int)
        the list of available indices.
    partition: float or int
        in case of float: the relative size of the share for which the indices to be returned
        in case of int: the number of elements to return starting from start_idx, [start_idx : start_idx + partition]
    start_idx: int
        the index to start at.
    shuffle_output: bool
        to shuffle the indices before being returned.

    Return:
    -------
    list, the list of indices

    """
    indices_arr = np.array(indices_arr)
    # if the partition is a percentage
    if isinstance(partition, float):
        size = int(np.round(partition * len(indices_arr)))
        indices = indices_arr[np.arange(start=start_idx, stop=start_idx + size, step=1)]

    # if the partition is the size of the share
    elif (isinstance(partition, int)) and (partition < len(indices_arr)):
        indices = indices_arr[np.arange(start=start_idx, stop=partition+start_idx)]
    else:
        raise  ValueError("partition should be either an integer or float")

    if shuffle_output:
        np.random.shuffle(indices)
    return indices

def save_dataset(dataset, indices, dst_path): #DOC OK
    """
    Params:
    -------
    dataset: torch.utils.data.Dataset
        the dataset to save all/part of it.
    indices: 
        when not None, images with these indices only are saved
    dst_path: 
        the path where to save the images
    
    Return
    ------
    void
    """
    indices = np.array(indices, dtype=int)
    imgs_names = np.array(dataset.imgs)

    # if the path does not exist, create it
    if not os.path.exists(dst_path):
        os.makedirs(dst_path, exist_ok=True)

    for file_path in imgs_names[indices]:
        shutil.copy(np.squeeze(file_path)[0].astype(str), dst_path)
        

def plot_confusion_matrix(y_true, y_pred, classes,
                              normalize=False,
                              title=None,
                              cmap=plt.cm.Blues):
    """
    # available on Scikit-learn  https://scikit-learn.org/stable/auto_examples/model_selection/plot_confusion_matrix.html
    This function prints and plots the confusion matrix.
    Normalization can be applied by setting `normalize=True`.
    """

    if not title:
        if normalize:
            title = 'Normalized confusion matrix'
        else:
            title = 'Confusion matrix, without normalization'

    # Compute confusion matrix
    cm = confusion_matrix(y_true, y_pred)
    # Only use the labels that appear in the data
    classes = classes[unique_labels(y_true, y_pred)]
    if normalize:
        cm = cm.astype('float') / cm.sum(axis=1)[:, np.newaxis]
        print("Normalized confusion matrix")
    else:
        print('Confusion matrix, without normalization')

    print(cm)

    fig, ax = plt.subplots()
    im = ax.imshow(cm, interpolation='nearest', cmap=cmap)
    ax.figure.colorbar(im, ax=ax)
    # We want to show all ticks...
    ax.set(xticks=np.arange(cm.shape[1]),
           yticks=np.arange(cm.shape[0]),
           # ... and label them with the respective list entries
           xticklabels=classes, yticklabels=classes,
           title=title,
           ylabel='True label',
           xlabel='Predicted label')

    # Rotate the tick labels and set their alignment.
    plt.setp(ax.get_xticklabels(), rotation=45, ha="right",
             rotation_mode="anchor")

    # Loop over data dimensions and create text annotations.
    fmt = '.2f' if normalize else 'd'
    thresh = cm.max() / 2.
    for i in range(cm.shape[0]):
        for j in range(cm.shape[1]):
            ax.text(j, i, format(cm[i, j], fmt),
                    ha="center", va="center",
                    color="white" if cm[i, j] > thresh else "black")
    fig.tight_layout()
    return ax

def write_xl(xlfile_name, data=None):
    """
    Params:
    -------
    xlfile_name: str
        the absolute path for the excel file to write inside
    data: iterable (list, tuple, dict, etc)
        the data to write.

    Returns:
    -------
        void
    """

    #create the spreadsheet if it doesnot exist
    if not os.path.exists(xlfile_name):
        book = Workbook()
    else:
        book = load_workbook(xlfile_name)

    # pick the active sheet
    sheet = book.active

    # write the data at the first non-empty line
    sheet.append(data)

    # save and close
    book.save(xlfile_name)
    book.close()

def dataset2matrix(folder_path):
    """
    This function is used to read a folder of images, flatten images and reshape them in a matrix as a sample every row.

    Params:
    -------
    folder_path: str
         the absolute path where the folder of images is located + the pattern.
         example: ~/home/user/Images/*.png
    Returns:
    --------
    labelled_samples: dict
        a ditionary containing two elements, samples and labels.
            'samples': np.ndarray
            The samples are a matrix where every row is a flattened sample.

            'labels': list(int)
            The labels are integers and start from 0 until number of subfolders in folder_path.
    """

    # initialization
    matrix, labels = [], []
    folders_names = []

    #split the pattern and path
    wild_card = os.path.split(folder_path)

    #walk through the folders
    for root, directory, file in os.walk(wild_card[0]):
        for folder in directory:
            folders_names.append(os.path.join(root, folder))

    index = 0
    # for all the folders
    for folder in folders_names:

        # read the names of the images in each folder
        files_names = glob(os.path.join(folder, wild_card[1]))

        # read all the images, reshape as rows, and stack them
        matrix.append(np.vstack([np.reshape(io.imread(file_name), (1,-1)) for file_name in files_names ]))

        # put a label for the images in each folder
        labels.append([index for file_name in files_names ])
        index+=1

    matrix = np.vstack(matrix)
    labels = np.hstack(labels)
    return({'samples':matrix, 'labels':labels})

    return matrix
